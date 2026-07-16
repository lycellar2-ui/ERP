import openpyxl
import os
import glob

def search_all_quotations():
    folder_path = r"D:\Lyscellar\Quotation"
    files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    keywords = ["Tổng", "VAT", "Thành tiền", "Ký", "Sign", "Bank", "Tài khoản", "Giám đốc", "Hummingbird", "Maii Bistro"]
    
    for file_path in files:
        basename = os.path.basename(file_path)
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            print(f"\nFile: {basename} | Sheets: {wb.sheetnames}")
            # Just read the first sheet or sheets with 'quote' or 'quotation' in their names
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Let's inspect first 100 rows
                for r in range(1, 101):
                    # In read-only mode, we can access cell values
                    try:
                        row_cells = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
                    except:
                        break
                    
                    for c_idx, val in enumerate(row_cells):
                        if val is not None:
                            val_str = str(val)
                            if any(k in val_str for k in keywords):
                                print(f"  [{sheet_name}] Row {r:02d} Col {c_idx+1}: {val_str[:50]}")
        except Exception as e:
            print(f"Error reading {basename}: {e}")

if __name__ == "__main__":
    search_all_quotations()
