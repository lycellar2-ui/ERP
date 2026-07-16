import openpyxl
import os

def inspect_same():
    file_path = r"D:\Lyscellar\Quotation\Ly's Cellars Quotation Template w pics and WS prices - Same.xlsx"
    if not os.path.exists(file_path):
        print("File not found")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
        for r in range(1, min(sheet.max_row + 1, 100)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 15))]
            if any(val is not None for val in row_vals):
                print(f"Row {r:02d}: {[str(v)[:50] if v is not None else '' for v in row_vals]}")

if __name__ == "__main__":
    inspect_same()
