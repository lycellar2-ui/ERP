import openpyxl
import os

def search_content():
    file_path = r"D:\Lyscellar\Quotation\Quotation - Hummingbird 16.06.xlsx"
    if not os.path.exists(file_path):
        print("File not found")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\nSheet: {name}")
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None and ("Hummingbird" in str(val) or "Báo giá" in str(val) or "Quotation" in str(val) or "Kính gửi" in str(val)):
                    print(f"  Cell({r},{c}): {val}")

if __name__ == "__main__":
    search_content()
