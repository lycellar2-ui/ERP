import openpyxl
import os

def inspect_xlsx():
    file_path = r"D:\Lyscellar\Quotation\Quotation - Hummingbird 16.06.xlsx"
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
        
        # Read the first 40 rows to see layout
        for r in range(1, min(sheet.max_row + 1, 60)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 15))]
            # Skip completely empty rows
            if any(val is not None for val in row_vals):
                formatted_vals = []
                for val in row_vals:
                    if val is None:
                        formatted_vals.append("")
                    elif isinstance(val, float):
                        formatted_vals.append(f"{val:.2f}")
                    else:
                        formatted_vals.append(str(val))
                # Truncate values for readability
                trimmed_vals = [val[:30] + "..." if len(val) > 30 else val for val in formatted_vals]
                print(f"Row {r:02d}: {trimmed_vals}")

if __name__ == "__main__":
    inspect_xlsx()
