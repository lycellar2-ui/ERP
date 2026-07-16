import openpyxl
import os

def inspect_styles():
    file_path = r"D:\Lyscellar\Quotation\Quotation - Hummingbird 16.06.xlsx"
    if not os.path.exists(file_path):
        print("File not found")
        return

    wb = openpyxl.load_workbook(file_path, data_only=False) # Load with formulas to see if there are any
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n================ Sheet: {name} ================")
        
        # Print top 15 rows with style details
        for r in range(1, 16):
            row_cells = []
            for c in range(1, min(sheet.max_column + 1, 12)):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                font = cell.font
                fill = cell.fill
                
                cell_info = ""
                if val is not None:
                    cell_info = f"'{val}'"
                
                # Check background color
                bg_color = None
                if fill and fill.start_color and fill.start_color.rgb:
                    bg_color = fill.start_color.rgb
                
                # Check font details
                font_info = []
                if font:
                    if font.bold: font_info.append("Bold")
                    if font.size: font_info.append(f"{font.size}pt")
                    if font.color and font.color.rgb: font_info.append(f"Color:{font.color.rgb}")
                
                style_desc = ""
                if bg_color or font_info:
                    style_desc = f" [BG:{bg_color} Font:{','.join(font_info)}]"
                
                if val is not None:
                    row_cells.append(f"Col {c}: {cell_info}{style_desc}")
            
            if row_cells:
                print(f"Row {r:02d} -> " + " | ".join(row_cells))

if __name__ == "__main__":
    inspect_styles()
