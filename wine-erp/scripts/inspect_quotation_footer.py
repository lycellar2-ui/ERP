import openpyxl
import os

def inspect_footers():
    file_path = r"D:\Lyscellar\Quotation\Quotation - Hummingbird 16.06.xlsx"
    if not os.path.exists(file_path):
        print("File not found")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n================ Sheet: {name} ================")
        start_row = max(1, sheet.max_row - 20)
        for r in range(start_row, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 15))]
            if any(val is not None for val in row_vals):
                print(f"Row {r:02d}: {[str(v)[:40] if v is not None else '' for v in row_vals]}")

if __name__ == "__main__":
    inspect_footers()
