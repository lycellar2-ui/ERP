import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Lyscellar\Kế toán\Kiểm kê\Kiểm kê TA 15.7.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheet names:", wb.sheetnames)

sheet = wb.active
rows = list(sheet.iter_rows(values_only=True))

print(f"Total rows in sheet: {len(rows)}")

for i, r in enumerate(rows[:20]):
    row_str = [str(cell) for cell in r if cell is not None]
    print(f"Row {i}: {row_str}")

out_data = []
for r in rows:
    out_data.append([str(c) if c is not None else "" for c in r])

with open(r"d:\Lyruou\wine-erp\prisma\inventory_raw.json", "w", encoding="utf-8") as f:
    json.dump(out_data, f, ensure_ascii=False, indent=2)

print("Saved raw inventory data to inventory_raw.json")
